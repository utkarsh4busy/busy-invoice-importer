import streamlit as st
import requests
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import json
from datetime import datetime
import time
import os
import uuid

# ─── SUPABASE ─────────────────────────────────────────────────────────────────
SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://zvziwaeeabfpdwqdektj.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "sb_publishable_pblalH1FTg5VLTYfikbv1w__ipChds0")

def sb_headers():
    return {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }

def db_insert(rows: list):
    try:
        r = requests.post(
            f"{SUPABASE_URL}/rest/v1/invoice_logs",
            headers=sb_headers(),
            json=rows,
            timeout=10,
        )
        r.raise_for_status()
    except Exception as e:
        st.warning(f"DB log failed: {e}")

def db_fetch_all() -> list:
    try:
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/invoice_logs?order=session_time.desc",
            headers=sb_headers(),
            timeout=10,
        )
        r.raise_for_status()
        return r.json()
    except Exception:
        return []

# ─── CONFIG ───────────────────────────────────────────────────────────────────
API_URL       = "https://ocr-preprod.busy.in/voucher/v1/ocr/"
DEFAULT_TOKEN = os.environ.get("OCR_TOKEN", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJhcHBsaWNhdGlvbl9pZCI6MiwiYXBwbGljYXRpb25fc3Vic2NyaXB0aW9uX2lkIjoyNywiZXhwIjoxNzc5NzczODQ5LCJzZXJ2aWNlIjoib2NyIn0.2xPWJe_P_qjSU9cq0dENSWF0ZqMLjUqySeHS5Wk3Xyk")
INVOICE_TYPE  = "purchase"

# ─── HARDCODED VENDOR MASTER ──────────────────────────────────────────────────
MASTER_VENDORS = {
    "Hazamin Consultancy (OPC) Pvt Ltd",
    "Gangwal Infrastructure Private Limited",
    "Eureka Coworking",
    "Ekatvam Infra Private Limited",
    "Collablabs Coworks Private Limited",
    "LANS FACICARE LLP",
    "Inspire Network Technology Solutions Pvt Ltd",
    "Innov8 Workspaces India Limited",
    "INDIAMART INTERMESH LIMITED",
    "IA India Accelerator Private Limited",
    "Segment Spaces and Infra Pvt Ltd",
    "91 Springboard Business Hub Private Limited",
    "Pink Hive Co-Working",
    "NEXT 57 Coworking",
    "Lavero Infra Services Private Limited",
    "LANS Facicare LLP",
    "Skillr Talent Private Limited",
    "VMS (referenced via stamp paper)",
    "Squadrun (referenced via stamp paper)",
    "Ocube (referenced via stamp paper)",
    "All Set (referenced via stamp paper)",
    "Germanium Technologies",
    "3I Business Solutions",
    "Interglobal HR Compliance",
    "AVI HR Software Pvt Ltd",
    "Prabhash Kumar Agarwal",
    "Indiamart Intermesh Limited",
    "Vivikt Growth LLP",
}

def check_vendor_match(party_name: str) -> tuple:
    """Exact string match against hardcoded vendor master."""
    name = (party_name or "").strip()
    if not name:
        return False, "Vendor name is empty in invoice"
    if name in MASTER_VENDORS:
        return True, ""
    return False, f"Vendor name mismatch: invoice has '{name}', not found in master agreement"

# ─── COLUMNS ──────────────────────────────────────────────────────────────────
COLUMNS = [
    "Invoice_Type", "Voucher_Date", "Voucher_No", "Supplier_Invoice_No",
    "Supplier_Invoice_Date", "Party_Name", "Party_GSTIN", "Party_PAN",
    "Party_Address", "Place_of_Supply", "Reverse_Charge", "Invoice_Category",
    "Expense_Category", "Employee_Name", "Department", "Cost_Center",
    "Item_HSN_SAC", "Item_Description", "Quantity", "Unit", "Rate_per_Unit",
    "Discount_Pct", "Taxable_Amount", "IGST_Pct", "IGST_Amount", "CGST_Pct",
    "CGST_Amount", "SGST_Pct", "SGST_Amount", "TCS_TDS_Amount",
    "Invoice_Total", "Payment_Terms", "Due_Date", "Narration"
]

# ─── API ──────────────────────────────────────────────────────────────────────
def call_ocr_api(file_bytes: bytes, filename: str, token: str):
    headers = {
        "Source": "web",
        "Service": "ocr",
        "Authorization": f"Bearer {token}",
    }
    files = {"file": (filename, file_bytes, "application/pdf")}
    data  = {"type": INVOICE_TYPE}
    token_preview = token[:12] + "..." + token[-6:] if len(token) > 20 else "***"
    request_meta = {
        "method": "POST",
        "url": API_URL,
        "headers": {
            "Source": "web",
            "Service": "ocr",
            "Authorization": f"Bearer {token_preview}",
            "Content-Type": "multipart/form-data",
        },
        "form_fields": {"type": INVOICE_TYPE},
        "file": {
            "field_name": "file",
            "filename": filename,
            "content_type": "application/pdf",
            "size_kb": round(len(file_bytes) / 1024, 1),
        },
    }
    resp = requests.post(API_URL, headers=headers, files=files, data=data, timeout=60)
    resp.raise_for_status()
    return resp.json(), request_meta

# ─── JSON → ROWS ──────────────────────────────────────────────────────────────
def determine_gst_type(seller_gstin: str, place_of_supply: str) -> str:
    if not seller_gstin or len(seller_gstin) < 2:
        return "igst"
    state_code = seller_gstin[:2]
    pos = place_of_supply or ""
    pos_code = ""
    if "(" in pos and ")" in pos:
        pos_code = pos[pos.index("(") + 1: pos.index(")")]
    return "cgst_sgst" if state_code == pos_code else "igst"

def build_party_address(seller: dict, buyer: dict) -> str:
    parts = []
    name  = seller.get("name") or ""
    addr  = seller.get("address") or ""
    if name:
        parts.append(name)
    if addr:
        parts.append(addr)
    buyer_addr = buyer.get("address") or ""
    pos = "Place of Supply : " + (buyer.get("city") or "")
    parts += [buyer_addr, pos]
    return ", ".join(p for p in parts if p)

def json_to_rows(api_json: dict) -> list:
    d          = api_json.get("data", {}).get("json_data", {})
    seller     = d.get("seller", {})
    buyer      = d.get("buyer", {})
    summary    = d.get("summary", {})
    line_items = d.get("line_items", [])
    place_of_supply = d.get("place_of_supply", "")
    pos_display = f"{place_of_supply} (07)" if place_of_supply and "(07)" not in place_of_supply else place_of_supply
    inv_no   = d.get("invoice_number", "")
    inv_date = d.get("invoice_date", "")
    grand_total    = summary.get("grand_total", 0)
    reverse_charge = d.get("reverse_charge", "N")
    gst_type       = determine_gst_type(seller.get("gstin", ""), pos_display)
    seller_gstin = seller.get("gstin", "")
    pan          = seller_gstin[2:12] if seller_gstin and len(seller_gstin) >= 12 else ""
    party_address = build_party_address(seller, buyer)
    narration     = f"Inv No: {inv_no} Dt: {inv_date}"

    def _num(val, default=0):
        try:
            return float(val) if val not in (None, "", "null") else default
        except (TypeError, ValueError):
            return default

    rows = []
    for item in line_items:
        rate     = _num(item.get("rate") or item.get("amount"), 0)
        qty      = _num(item.get("quantity"), 1) or 1
        gst_rate = _num(item.get("gst_rate"), 0)
        taxable  = rate * qty
        igst_amt = cgst_amt = sgst_amt = 0.0
        igst_pct = cgst_pct = sgst_pct = 0.0
        if gst_type == "igst":
            igst_amt = round(taxable * gst_rate / 100, 2)
            igst_pct = gst_rate
        else:
            half     = gst_rate / 2
            cgst_amt = sgst_amt = round(taxable * half / 100, 2)
            cgst_pct = sgst_pct = half
        tcs_tds = round(taxable * 0.02)
        row = {
            "Invoice_Type":          "Purchase",
            "Voucher_Date":          inv_date,
            "Voucher_No":            "",
            "Supplier_Invoice_No":   inv_no,
            "Supplier_Invoice_Date": inv_date,
            "Party_Name":            seller.get("name") or seller.get("trade_name") or "",
            "Party_GSTIN":           seller_gstin,
            "Party_PAN":             pan,
            "Party_Address":         party_address,
            "Place_of_Supply":       pos_display,
            "Reverse_Charge":        reverse_charge,
            "Invoice_Category":      "Services",
            "Expense_Category":      "",
            "Employee_Name":         "",
            "Department":            "",
            "Cost_Center":           "",
            "Item_HSN_SAC":          item.get("hsn_sac") or "",
            "Item_Description":      item.get("description") or "",
            "Quantity":              qty,
            "Unit":                  item.get("unit") or "Pcs.",
            "Rate_per_Unit":         rate,
            "Discount_Pct":          0.00,
            "Taxable_Amount":        taxable,
            "IGST_Pct":              igst_pct,
            "IGST_Amount":           igst_amt,
            "CGST_Pct":              cgst_pct,
            "CGST_Amount":           cgst_amt,
            "SGST_Pct":              sgst_pct,
            "SGST_Amount":           sgst_amt,
            "TCS_TDS_Amount":        tcs_tds,
            "Invoice_Total":         grand_total,
            "Payment_Terms":         "",
            "Due_Date":              "",
            "Narration":             narration,
        }
        rows.append(row)
    return rows

# ─── EXCEL BUILDER ────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=9)
TITLE_FONT  = Font(name="Arial", bold=True, size=11, color="1F4E79")
DATA_FONT   = Font(name="Arial", size=9)
ALT_FILL    = PatternFill("solid", start_color="EBF3FB")
BORDER_SIDE = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)

COL_WIDTHS = {
    "Invoice_Type": 12, "Voucher_Date": 13, "Voucher_No": 11,
    "Supplier_Invoice_No": 20, "Supplier_Invoice_Date": 20,
    "Party_Name": 28, "Party_GSTIN": 18, "Party_PAN": 13,
    "Party_Address": 40, "Place_of_Supply": 16, "Reverse_Charge": 14,
    "Invoice_Category": 16, "Expense_Category": 16, "Employee_Name": 14,
    "Department": 13, "Cost_Center": 12, "Item_HSN_SAC": 13,
    "Item_Description": 35, "Quantity": 9, "Unit": 7,
    "Rate_per_Unit": 13, "Discount_Pct": 12, "Taxable_Amount": 14,
    "IGST_Pct": 9, "IGST_Amount": 12, "CGST_Pct": 9, "CGST_Amount": 12,
    "SGST_Pct": 9, "SGST_Amount": 12, "TCS_TDS_Amount": 14,
    "Invoice_Total": 13, "Payment_Terms": 13, "Due_Date": 11, "Narration": 38,
}

def build_excel(all_rows: list, generated_at: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Busy_Import_Data"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    title_cell = ws.cell(row=1, column=1, value=f"BUSY INVOICE IMPORT DATA — Generated {generated_at}")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 18
    for ci, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=ci, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 28
    for ri, row in enumerate(all_rows, start=3):
        fill = ALT_FILL if ri % 2 == 0 else None
        for ci, col in enumerate(COLUMNS, start=1):
            val  = row.get(col, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col == "Party_Address"))
            if fill:
                cell.fill = fill
    for ci, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 12)
    ws.freeze_panes = "A3"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ─── STREAMLIT CONFIG ─────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Busy Invoice Importer",
    page_icon="📄",
    layout="wide",
)

st.markdown("""
<style>
    .main { background: #F0F4FA; }
    [data-testid="stAppViewContainer"] { background: #F0F4FA; }
    [data-testid="stMain"] { background: #F0F4FA; }
    .stButton > button {
        background: #1F4E79; color: white; border-radius: 6px;
        font-weight: 600; padding: 0.5rem 1.5rem;
    }
    .stButton > button:hover { background: #2E75B6; }
    .metric-card {
        background: white; border-radius: 10px; padding: 1rem 1.5rem;
        box-shadow: 0 1px 4px rgba(0,0,0,0.08); text-align: center;
    }
    .metric-card h2 { color: #1F4E79; margin: 0; font-size: 2rem; }
    .metric-card p  { color: #666; margin: 0; font-size: 0.85rem; }
    .status-box {
        border-radius: 8px; padding: 0.6rem 1rem; margin: 4px 0;
        font-size: 0.87rem;
    }
    .success   { background: #E8F5E9; border-left: 4px solid #4CAF50; color: #2E7D32; }
    .error     { background: #FFEBEE; border-left: 4px solid #F44336; color: #C62828; }
    .matched   { background: #E3F2FD; border-left: 4px solid #1565C0; color: #0D47A1; }
    .unmatched { background: #FFF3E0; border-left: 4px solid #E65100; color: #BF360C; }
    .timing-row  { display: flex; gap: 8px; flex-wrap: wrap; margin-top: 6px; }
    .timing-pill { background: #F0FDF4; border: 1px solid #BBF7D0; border-radius: 20px;
        padding: 3px 10px; font-size: 0.8rem; color: #166534; font-weight: 500; }
    .timing-pill.ocr   { background: #EFF6FF; border-color: #BFDBFE; color: #1E40AF; }
    .timing-pill.excel { background: #FFF7ED; border-color: #FED7AA; color: #9A3412; }
    .dash-table { width: 100%; border-collapse: collapse; font-size: 0.85rem; }
    .dash-table th { background: #1F4E79; color: white; padding: 8px 12px; text-align: left; }
    .dash-table td { padding: 7px 12px; border-bottom: 1px solid #E5E7EB; }
    .dash-table tr:hover td { background: #F0F4FA; }
    .badge-success  { background:#E8F5E9; color:#2E7D32; border-radius:12px; padding:2px 10px; font-size:0.78rem; font-weight:600; }
    .badge-failed   { background:#FFEBEE; color:#C62828; border-radius:12px; padding:2px 10px; font-size:0.78rem; font-weight:600; }
    .badge-matched  { background:#E3F2FD; color:#0D47A1; border-radius:12px; padding:2px 10px; font-size:0.78rem; font-weight:600; }
    .badge-unmatched{ background:#FFF3E0; color:#BF360C; border-radius:12px; padding:2px 10px; font-size:0.78rem; font-weight:600; }
    .badge-na       { background:#F3F4F6; color:#6B7280; border-radius:12px; padding:2px 10px; font-size:0.78rem; font-weight:600; }
</style>
""", unsafe_allow_html=True)

# ─── HEADER ───────────────────────────────────────────────────────────────────
st.markdown("## 📄 Busy Invoice Importer")
st.markdown("Upload purchase PDFs → OCR API → Excel for Busy import.")
st.divider()

# ─── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Settings")
    token = st.text_input("Bearer Token", value=DEFAULT_TOKEN, type="password")
    st.markdown("---")
    st.markdown("**API Endpoint**")
    st.code(API_URL, language=None)
    st.markdown("**Invoice Type**")
    st.code(INVOICE_TYPE, language=None)
    st.markdown("---")
    st.markdown("**Vendor Master**")
    st.caption(f"{len(MASTER_VENDORS)} vendors loaded")
    with st.expander("View vendors"):
        for v in sorted(MASTER_VENDORS):
            st.markdown(f"- {v}")
    st.markdown("---")
    st.markdown("**How it works**")
    st.markdown("1. Upload one or more PDFs\n2. Each PDF is sent to the OCR API\n3. Responses are mapped to Busy import format\n4. Results logged to database\n5. Download the filled Excel")

# ─── SESSION STATE ────────────────────────────────────────────────────────────
if "results" not in st.session_state:
    st.session_state.results = []
if "excel_bytes" not in st.session_state:
    st.session_state.excel_bytes = None

# ─── TABS ─────────────────────────────────────────────────────────────────────
tab_dash, tab_upload, tab_results = st.tabs(["📊 Dashboard", "📤 Upload & Process", "📋 Results & Download"])

# ─── TAB 0: DASHBOARD ─────────────────────────────────────────────────────────
with tab_dash:
    st.markdown("### Upload History")
    if st.button("🔄 Refresh", key="refresh_dash"):
        st.rerun()

    logs = db_fetch_all()
    if not logs:
        st.info("No uploads logged yet. Process some invoices first.")
    else:
        df_logs = pd.DataFrame(logs)
        total     = len(df_logs)
        success   = int((df_logs["ocr_status"] == "success").sum())
        failed    = total - success
        matched   = int((df_logs["match_status"] == "matched").sum())
        unmatched = int((df_logs["match_status"] == "unmatched").sum())

        c1, c2, c3, c4, c5 = st.columns(5)
        with c1:
            st.markdown(f'<div class="metric-card"><h2>{total}</h2><p>Total Invoices</p></div>', unsafe_allow_html=True)
        with c2:
            st.markdown(f'<div class="metric-card"><h2 style="color:#4CAF50">{success}</h2><p>OCR Success</p></div>', unsafe_allow_html=True)
        with c3:
            st.markdown(f'<div class="metric-card"><h2 style="color:#F44336">{failed}</h2><p>OCR Failed</p></div>', unsafe_allow_html=True)
        with c4:
            st.markdown(f'<div class="metric-card"><h2 style="color:#1565C0">{matched}</h2><p>Vendor Matched</p></div>', unsafe_allow_html=True)
        with c5:
            st.markdown(f'<div class="metric-card"><h2 style="color:#E65100">{unmatched}</h2><p>Vendor Unmatched</p></div>', unsafe_allow_html=True)

        st.markdown("")
        sessions = df_logs.groupby("session_id", sort=False)
        for session_id, group in sessions:
            session_time = group["session_time"].iloc[0]
            n      = len(group)
            s_ok   = int((group["ocr_status"] == "success").sum())
            s_fail = n - s_ok
            label  = f"🗂 Session: {session_time}  ·  {n} invoice(s)  ·  {s_ok} ✅  {s_fail} ❌"
            with st.expander(label, expanded=False):
                rows_html = ""
                for _, row in group.iterrows():
                    ocr_badge = '<span class="badge-success">Success</span>' if row["ocr_status"] == "success" else '<span class="badge-failed">Failed</span>'
                    if row["match_status"] == "matched":
                        match_badge = '<span class="badge-matched">Matched</span>'
                    elif row["match_status"] == "unmatched":
                        match_badge = '<span class="badge-unmatched">Unmatched</span>'
                    else:
                        match_badge = '<span class="badge-na">N/A</span>'
                    mismatch_cell = row.get("mismatch_detail") or "—"
                    inv_no = row.get("invoice_number") or "—"
                    rows_html += f"""<tr>
                        <td>{row['filename']}</td>
                        <td>{inv_no}</td>
                        <td>{ocr_badge}</td>
                        <td>{match_badge}</td>
                        <td style="color:#666;font-size:0.8rem">{mismatch_cell}</td>
                    </tr>"""
                st.markdown(f"""
                <table class="dash-table"><thead><tr>
                    <th>Filename</th><th>Invoice No.</th><th>OCR Status</th>
                    <th>Vendor Match</th><th>Mismatch Detail</th>
                </tr></thead><tbody>{rows_html}</tbody></table>
                """, unsafe_allow_html=True)

# ─── TAB 1: UPLOAD ────────────────────────────────────────────────────────────
with tab_upload:
    col_upload, col_info = st.columns([2, 1])
    with col_upload:
        uploaded_files = st.file_uploader(
            "Drop PDF invoices here (single or bulk)",
            type=["pdf"],
            accept_multiple_files=True,
        )
    with col_info:
        if uploaded_files:
            st.info(f"**{len(uploaded_files)} file(s)** ready to process")
            for f in uploaded_files:
                sz = len(f.getvalue()) / 1024
                st.markdown(f"- `{f.name}` ({sz:.1f} KB)")

    if uploaded_files:
        st.markdown("")
        process_btn = st.button("🚀 Process All PDFs", type="primary")
        if process_btn:
            st.session_state.results    = []
            st.session_state.excel_bytes = None
            all_rows     = []
            progress     = st.progress(0, text="Starting...")
            log_area     = st.container()
            session_id   = str(uuid.uuid4())
            session_time = datetime.utcnow().isoformat()
            db_rows      = []

            for i, pdf_file in enumerate(uploaded_files):
                pct = int((i / len(uploaded_files)) * 100)
                progress.progress(pct, text=f"Processing {pdf_file.name} ({i+1}/{len(uploaded_files)})…")
                file_bytes = pdf_file.getvalue()
                result = {
                    "filename": pdf_file.name, "status": None, "rows": [],
                    "error": "", "raw_json": None, "ocr_time": None,
                    "excel_time": None, "request_meta": None,
                    "match": None, "mismatch_detail": "", "invoice_number": "",
                }
                try:
                    t0 = time.perf_counter()
                    raw, req_meta = call_ocr_api(file_bytes, pdf_file.name, token)
                    t1 = time.perf_counter()
                    result["ocr_time"]     = t1 - t0
                    result["raw_json"]     = raw
                    result["request_meta"] = req_meta

                    if raw.get("status"):
                        rows = json_to_rows(raw)
                        t2   = time.perf_counter()
                        result["excel_time"] = t2 - t1
                        result["rows"]       = rows
                        result["status"]     = "success"
                        all_rows.extend(rows)

                        inv_no     = raw.get("data", {}).get("json_data", {}).get("invoice_number") or ""
                        party_name = rows[0]["Party_Name"] if rows else ""
                        result["invoice_number"] = inv_no

                        matched, mismatch = check_vendor_match(party_name)
                        result["match"]           = matched
                        result["mismatch_detail"] = mismatch

                        match_line = (
                            f'<br>🏢 Vendor: <b>{party_name}</b> — '
                            + ('✅ <span style="color:#2E7D32">Matched in master agreement</span>'
                               if matched else
                               f'⚠️ <span style="color:#E65100">{mismatch}</span>')
                        )
                        with log_area:
                            st.markdown(
                                f'<div class="status-box success">✅ <b>{pdf_file.name}</b> — {len(rows)} line item(s) extracted'
                                f'{match_line}'
                                f'<div class="timing-row">'
                                f'<span class="timing-pill ocr">🔍 OCR: {result["ocr_time"]:.2f}s</span>'
                                f'<span class="timing-pill excel">📊 Excel mapping: {result["excel_time"]:.3f}s</span>'
                                f'</div></div>',
                                unsafe_allow_html=True,
                            )
                        db_rows.append({
                            "session_id":      session_id,
                            "session_time":    session_time,
                            "filename":        pdf_file.name,
                            "invoice_number":  inv_no,
                            "ocr_status":      "success",
                            "match_status":    "matched" if matched else "unmatched",
                            "mismatch_detail": mismatch or None,
                        })
                    else:
                        result["status"] = "error"
                        result["error"]  = raw.get("message", "Unknown API error")
                        with log_area:
                            st.markdown(
                                f'<div class="status-box error">❌ <b>{pdf_file.name}</b> — {result["error"]}</div>',
                                unsafe_allow_html=True,
                            )
                        db_rows.append({
                            "session_id":      session_id,
                            "session_time":    session_time,
                            "filename":        pdf_file.name,
                            "invoice_number":  None,
                            "ocr_status":      "failed",
                            "match_status":    "n/a",
                            "mismatch_detail": result["error"],
                        })
                except Exception as e:
                    result["status"] = "error"
                    result["error"]  = str(e)
                    with log_area:
                        st.markdown(
                            f'<div class="status-box error">❌ <b>{pdf_file.name}</b> — {e}</div>',
                            unsafe_allow_html=True,
                        )
                    db_rows.append({
                        "session_id":      session_id,
                        "session_time":    session_time,
                        "filename":        pdf_file.name,
                        "invoice_number":  None,
                        "ocr_status":      "failed",
                        "match_status":    "n/a",
                        "mismatch_detail": str(e),
                    })
                st.session_state.results.append(result)

            progress.progress(100, text="Done!")
            if db_rows:
                db_insert(db_rows)
            if all_rows:
                ts = datetime.now().strftime("%d-%m-%Y %H:%M")
                st.session_state.excel_bytes = build_excel(all_rows, ts)
                success_count = sum(1 for r in st.session_state.results if r["status"] == "success")
                fail_count    = len(st.session_state.results) - success_count
                st.success(f"✅ Done — {success_count} succeeded, {fail_count} failed. Go to **Results & Download** tab.")
            else:
                st.error("No rows extracted. Check API token or PDF contents.")

# ─── TAB 2: RESULTS ───────────────────────────────────────────────────────────
with tab_results:
    if not st.session_state.results:
        st.info("No results yet. Upload PDFs and click **Process All PDFs**.")
    else:
        results       = st.session_state.results
        success_count = sum(1 for r in results if r["status"] == "success")
        fail_count    = len(results) - success_count
        total_rows    = sum(len(r["rows"]) for r in results)

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.markdown(f'<div class="metric-card"><h2>{len(results)}</h2><p>PDFs Processed</p></div>', unsafe_allow_html=True)
        with c2:
            st.markdown(f'<div class="metric-card"><h2 style="color:#4CAF50">{success_count}</h2><p>Succeeded</p></div>', unsafe_allow_html=True)
        with c3:
            st.markdown(f'<div class="metric-card"><h2 style="color:#F44336">{fail_count}</h2><p>Failed</p></div>', unsafe_allow_html=True)
        with c4:
            st.markdown(f'<div class="metric-card"><h2>{total_rows}</h2><p>Line Items</p></div>', unsafe_allow_html=True)

        st.markdown("")
        if st.session_state.excel_bytes:
            fname = f"Busy_Import_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            st.download_button(
                label="⬇️ Download Excel (Busy Import Format)",
                data=st.session_state.excel_bytes,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )
            st.markdown("")

        st.markdown("### Per-File Details")
        for r in results:
            status_icon = "✅" if r["status"] == "success" else "❌"
            match_icon  = "" if r.get("match") is None else ("🟢" if r["match"] else "🟠")
            with st.expander(f"{status_icon} {match_icon} {r['filename']} — {len(r['rows'])} rows"):
                ocr_t   = r.get("ocr_time")
                excel_t = r.get("excel_time")
                if ocr_t or excel_t:
                    total_t = (ocr_t or 0) + (excel_t or 0)
                    pills   = ""
                    if ocr_t:
                        pills += f'<span class="timing-pill ocr">🔍 OCR: {ocr_t:.2f}s</span>'
                    if excel_t:
                        pills += f'<span class="timing-pill excel">📊 Excel mapping: {excel_t:.3f}s</span>'
                    pills += f'<span class="timing-pill">⏱ Total: {total_t:.2f}s</span>'
                    st.markdown(f'<div class="timing-row" style="margin-bottom:10px">{pills}</div>', unsafe_allow_html=True)

                if r["status"] == "success" and r["rows"]:
                    if r.get("match") is True:
                        st.markdown('<div class="status-box matched">🟢 Vendor matched in master agreement</div>', unsafe_allow_html=True)
                    elif r.get("match") is False:
                        st.markdown(f'<div class="status-box unmatched">🟠 {r["mismatch_detail"]}</div>', unsafe_allow_html=True)
                    df_preview = pd.DataFrame(r["rows"])
                    st.dataframe(
                        df_preview[["Supplier_Invoice_No", "Party_Name", "Item_Description",
                                    "Rate_per_Unit", "IGST_Pct", "IGST_Amount", "Invoice_Total"]],
                        use_container_width=True,
                    )
                else:
                    st.error(f"Error: {r['error']}")

                if r.get("request_meta") or r.get("raw_json"):
                    st.markdown("#### 🔌 API Inspector")
                    api_req_tab, api_resp_tab = st.tabs(["📤 Request", "📥 Response"])
                    with api_req_tab:
                        if r.get("request_meta"):
                            m = r["request_meta"]
                            st.markdown(f"**`{m['method']}`** → `{m['url']}`")
                            st.json(m["headers"])
                            col_ff, col_fi = st.columns(2)
                            with col_ff:
                                st.json(m["form_fields"])
                            with col_fi:
                                st.json(m["file"])
                    with api_resp_tab:
                        if r.get("raw_json"):
                            resp = r["raw_json"]
                            st.markdown(f"**{'✅ status: true' if resp.get('status') else '❌ status: false'}**")
                            st.json(resp)